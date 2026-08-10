#!/usr/bin/env python3
"""Build plot-ready summary rows directly from saved study caches."""

from __future__ import annotations

import argparse
from copy import copy
import json
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT))
import meta_analysis  # noqa: E402


DISEASES = {
    "breast": "Breast cancer",
    "ovarian": "Ovarian cancer",
    "uterine": "Uterine cancer",
}

COLUMNS = [
    "Exposure",
    "number studies",
    "Pooled RR",
    "lower CI RR",
    "upper CI RR",
    "lower PI RR",
    "upper PI RR",
    "I^2 (%)",
    "eggers p-value",
    "total N",
    "total Cases",
]

# Keep the paper-plot selection in sync with the browser UI. When an article
# reports its cohort size but not its case count, MetaFemina estimates cases
# from the disease-specific lifetime risk before applying the default >50-case
# threshold (see static/script.js:renderStudiesTable).
DISEASE_PREVALENCE = {
    "Breast cancer": 0.13,
    "Ovarian cancer": 0.013,
    "Uterine cancer": 0.031,
}


def analysis_case_counts(frame: pd.DataFrame, disease: str) -> pd.Series:
    """Return reported cases, saved estimates, or UI-equivalent estimates."""
    reported = frame.get("Cases", pd.Series(np.nan, index=frame.index))
    saved_estimate = frame.get("Estimated Cases", pd.Series(np.nan, index=frame.index))
    sample_size = frame.get("Sample Size", pd.Series(np.nan, index=frame.index))
    prevalence = DISEASE_PREVALENCE[disease]

    # JavaScript Math.round is floor(x + 0.5) for non-negative sample sizes;
    # pandas Series.round uses bankers' rounding and can differ at .5.
    derived_estimate = np.floor(sample_size * prevalence + 0.5)
    return reported.fillna(saved_estimate).fillna(derived_estimate)


def summarize_cache(path: Path, disease: str, exposure: str, dietary: bool) -> dict | None:
    payload = json.loads(path.read_text(encoding="utf-8"))
    frame = pd.DataFrame(payload.get("studies", []))
    if frame.empty:
        return None
    frame = meta_analysis.filter_curated_meta_analysis_exclusions(
        frame, disease, exposure, "Incidence"
    )
    if frame.empty:
        return None
    if dietary:
        if "exposure_measurement_type" not in frame:
            return None
        measurement = frame["exposure_measurement_type"].fillna("").astype(str).str.strip().str.lower()
        frame = frame.loc[measurement.eq("dietary_intake")].copy()
    if frame.empty:
        return None
    for column in ("Effect Size", "Lower CI", "Upper CI", "Cases", "Sample Size", "Estimated Cases"):
        if column in frame:
            frame[column] = pd.to_numeric(
                frame[column].astype(str).str.replace(",", "", regex=False).str.strip(),
                errors="coerce",
            )
    cases = analysis_case_counts(frame, disease)
    eligible_effect = frame.get("Effect Type", pd.Series("", index=frame.index)).map(
        meta_analysis.is_eligible_effect_type
    )
    quality = frame.get("Quality Score", pd.Series("Fair", index=frame.index)).fillna("Fair").astype(str).str.strip().str.lower()
    valid = frame.loc[
        (frame["Effect Size"] > 0)
        & (frame["Lower CI"] > 0)
        & (frame["Upper CI"] > 0)
        & (cases > 50)
        & eligible_effect
        & quality.isin({"good", "moderate"})
    ].copy()
    if valid.empty:
        return None
    result = meta_analysis.perform_meta_analysis(
        valid, disease, exposure, generate_plots=False, df_all=valid
    )
    headline = result.get("headline")
    if not headline:
        return None
    sample_sizes = valid.get("Sample Size", pd.Series(0, index=valid.index)).fillna(0)
    case_counts = cases.loc[valid.index].fillna(0)
    return {
        "Exposure": exposure,
        "number studies": int(len(valid)),
        "Pooled RR": headline.get("pooled_es"),
        "lower CI RR": headline.get("ci_low"),
        "upper CI RR": headline.get("ci_upp"),
        "lower PI RR": headline.get("pi_low"),
        "upper PI RR": headline.get("pi_upp"),
        "I^2 (%)": round(float(headline.get("i2") or 0), 1),
        "eggers p-value": headline.get("eggers_p"),
        "total N": int(sample_sizes.sum()),
        "total Cases": int(case_counts.sum()),
    }


def build() -> dict:
    output = {"combined": {}, "dietary": {}}
    cache_root = REPO_ROOT / "Cached_results"
    for disease_key, disease in DISEASES.items():
        filename = f"{disease.replace(' ', '_').lower()}_incidence_true_all.json"
        for dataset, dietary in (("combined", False), ("dietary", True)):
            rows = []
            for path in sorted(cache_root.glob(f"*/{filename}")):
                if path.parent.name == "multivitamin":
                    continue
                try:
                    row = summarize_cache(path, disease, path.parent.name, dietary)
                except Exception as exc:
                    print(f"Skipped {path}: {exc}", file=sys.stderr)
                    continue
                if row:
                    rows.append(row)
            output[dataset][disease_key] = sorted(rows, key=lambda row: row["Pooled RR"])
    return output


def sync_workbook(path: Path, rows: list[dict]) -> None:
    """Replace workbook data while preserving its established presentation."""
    workbook = load_workbook(path)
    worksheet = workbook[workbook.sheetnames[0]]
    headers = [cell.value for cell in worksheet[1]]
    if headers != COLUMNS:
        raise RuntimeError(f"Unexpected workbook columns in {path}: {headers}")

    old_data_rows = max(worksheet.max_row - 1, 0)
    template_row = 2 if old_data_rows else None
    for row_number, row in enumerate(rows, start=2):
        for column_number, header in enumerate(COLUMNS, start=1):
            cell = worksheet.cell(row_number, column_number)
            if template_row is not None and row_number > old_data_rows + 1:
                template = worksheet.cell(template_row, column_number)
                cell._style = copy(template._style)
                if template.has_style:
                    cell.number_format = template.number_format
                if template.alignment:
                    cell.alignment = copy(template.alignment)
            value = row.get(header)
            cell.value = None if pd.isna(value) else value

    new_last_row = len(rows) + 1
    if worksheet.max_row > new_last_row:
        worksheet.delete_rows(new_last_row + 1, worksheet.max_row - new_last_row)
    worksheet.auto_filter.ref = f"A1:K{new_last_row}"
    workbook.save(path)


def sync_workbooks(data: dict, workbook_dir: Path) -> None:
    for dataset in ("combined", "dietary"):
        for disease_key in DISEASES:
            path = workbook_dir / f"exposures_meta_analysis_{disease_key}_{dataset}.xlsx"
            if not path.exists():
                raise FileNotFoundError(f"Required plot workbook is missing: {path}")
            sync_workbook(path, data[dataset][disease_key])


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument(
        "--workbook-dir",
        type=Path,
        help="Also synchronize the six existing Plot workbooks in this directory.",
    )
    args = parser.parse_args()
    data = build()
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(data, indent=2) + "\n", encoding="utf-8")
    if args.workbook_dir:
        sync_workbooks(data, args.workbook_dir)
    print(args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
