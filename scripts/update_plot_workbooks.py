#!/usr/bin/env python3
"""Update paper workbooks only for exposure-cancer analyses changed monthly."""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT))
from scripts.build_plot_analysis_json import summarize_cache  # noqa: E402


WORKBOOKS = {
    "Breast cancer": REPO_ROOT / "Plot" / "exposures_meta_analysis_breast_combined.xlsx",
    "Ovarian cancer": REPO_ROOT / "Plot" / "exposures_meta_analysis_ovarian_combined.xlsx",
    "Uterine cancer": REPO_ROOT / "Plot" / "exposures_meta_analysis_uterine_combined.xlsx",
}

COLUMNS = [
    "Exposure",
    "number studies",
    "Pooled RR",
    "lower CI RR",
    "upper CI RR",
    "lower PI RR",
    "upper PI RR",
    "I^2 (%)",
    "eggers p-value",
    "total N",
    "total Cases",
]


def safe_name(value: str) -> str:
    return re.sub(r"[^a-zA-Z0-9]+", "_", value.lower()).strip("_")


def filtered_result(exposure: str, cancer: str) -> dict:
    cache_path = (
        REPO_ROOT
        / "Cached_results"
        / safe_name(exposure)
        / f"{safe_name(cancer)}_incidence_true_all.json"
    )
    result = summarize_cache(cache_path, cancer, safe_name(exposure), dietary=False)
    if not result:
        raise RuntimeError(f"Paper meta-analysis failed for {exposure}/{cancer}")
    return result


def update_workbook(path: Path, row: dict) -> None:
    workbook = load_workbook(path)
    worksheet = workbook[workbook.sheetnames[0]]
    headers = {str(cell.value): cell.column for cell in worksheet[1]}
    if list(headers) != COLUMNS:
        raise RuntimeError(f"Unexpected workbook columns in {path}: {list(headers)}")
    exposure_column = headers["Exposure"]
    target_row = None
    for row_number in range(2, worksheet.max_row + 1):
        if str(worksheet.cell(row_number, exposure_column).value).casefold() == row["Exposure"].casefold():
            target_row = row_number
            break
    if target_row is None:
        raise RuntimeError(f"Exposure {row['Exposure']} is absent from {path}")
    for header in COLUMNS:
        worksheet.cell(target_row, headers[header]).value = row[header]
    workbook.save(path)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--applied-changes", required=True, type=Path)
    args = parser.parse_args()
    with args.applied_changes.open(encoding="utf-8") as handle:
        changes = json.load(handle).get("changes", [])
    updated = []
    for change in changes:
        if change.get("status") != "updated":
            continue
        exposure = change["exposure"]
        cancer = change["cancer"]
        if cancer not in WORKBOOKS:
            continue
        row = filtered_result(exposure, cancer)
        update_workbook(WORKBOOKS[cancer], row)
        updated.append({"workbook": str(WORKBOOKS[cancer]), "row": row})
    print(json.dumps({"updated": updated}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
