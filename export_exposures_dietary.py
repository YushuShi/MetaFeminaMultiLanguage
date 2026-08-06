import os
import json
import pandas as pd
import numpy as np
import warnings
import meta_analysis
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Suppress runtime warnings from meta-analysis
warnings.filterwarnings('ignore')

DIETARY_MEASUREMENT_TYPE = "dietary_intake"

WORKBOOK_COLUMN_WIDTHS = {
    "Exposure": 28,
    "number studies": 15,
    "Pooled RR": 13,
    "lower CI RR": 14,
    "upper CI RR": 14,
    "lower PI RR": 14,
    "upper PI RR": 14,
    "I^2 (%)": 12,
    "eggers p-value": 16,
    "total N": 15,
    "total Cases": 15,
}

DISEASES = {
    "breast": {
        "file_pattern": "breast_cancer_incidence_true_all.json",
        "disease_label": "Breast Cancer",
        "context_template": "breast_cancer_{}_incidence",
        "output_file": "exposures_meta_analysis_breast_dietary.xlsx",
    },
    "ovarian": {
        "file_pattern": "ovarian_cancer_incidence_true_all.json",
        "disease_label": "Ovarian Cancer",
        "context_template": "ovarian_cancer_{}_incidence",
        "output_file": "exposures_meta_analysis_ovarian_dietary.xlsx",
    },
    "uterine": {
        "file_pattern": "uterine_cancer_incidence_true_all.json",
        "disease_label": "Uterine Cancer",
        "context_template": "uterine_cancer_{}_incidence",
        "output_file": "exposures_meta_analysis_uterine_dietary.xlsx",
    },
}


def format_workbook(path):
    workbook = load_workbook(path)
    worksheet = workbook.active
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False
    worksheet.row_dimensions[1].height = 30

    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_font = Font(color="FFFFFF", bold=True)
    headers = {cell.value: cell.column for cell in worksheet[1]}

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for header, column_index in headers.items():
        worksheet.column_dimensions[worksheet.cell(1, column_index).column_letter].width = (
            WORKBOOK_COLUMN_WIDTHS.get(header, 14)
        )

    count_headers = {"number studies", "total N", "total Cases"}
    one_decimal_headers = {"I^2 (%)"}
    four_decimal_headers = {"eggers p-value"}
    for header, column_index in headers.items():
        for row_index in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row_index, column_index)
            if header in count_headers:
                cell.number_format = "#,##0"
            elif header in one_decimal_headers:
                cell.number_format = "0.0"
            elif header in four_decimal_headers:
                cell.number_format = "0.0000"
            elif header != "Exposure":
                cell.number_format = "0.00"

    workbook.save(path)

def run_export(disease_key, disease_config):
    results_dir = 'Cached_results'
    all_results = []

    if not os.path.exists(results_dir):
        print(f"Directory {results_dir} not found.")
        return

    folders = os.listdir(results_dir)
    print(f"\n{'='*60}")
    print(f"  {disease_config['disease_label']}  —  {len(folders)} exposure folders")
    print(f"{'='*60}")

    blacklist = ['multivitamin']

    for folder in folders:
        if folder in blacklist:
            continue

        file_path = os.path.join(results_dir, folder, disease_config['file_pattern'])
        if not os.path.exists(file_path):
            continue

        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            studies = data.get('studies', [])
            if not studies:
                continue

            # Crowdsourced reports are advisory only and do not alter exports.
            cleaned_studies = [dict(s, exclusions=0) for s in studies]

            if not cleaned_studies:
                continue

            df = pd.DataFrame(cleaned_studies)

            # Dietary-only analysis: exclude human biospecimens, unclear
            # measurement types, and any future non-dietary classifications.
            if 'exposure_measurement_type' not in df.columns:
                continue
            measurement_type = (
                df['exposure_measurement_type']
                .fillna('')
                .astype(str)
                .str.strip()
                .str.lower()
            )
            df = df[measurement_type.eq(DIETARY_MEASUREMENT_TYPE)].copy()

            if len(df) == 0:
                continue

            # Numeric conversion
            for col in ['Effect Size', 'Lower CI', 'Upper CI', 'Cases', 'Sample Size', 'Estimated Cases']:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', '').str.strip(), errors='coerce')

            # Filter Cases >= 50 (standardised)
            cases_col = df['Cases'].fillna(df.get('Estimated Cases', np.nan)) if 'Cases' in df.columns else df.get('Estimated Cases', np.nan)
            effect_types = df.get('Effect Type', pd.Series('', index=df.index)).fillna('').astype(str).str.strip().str.upper()
            quality_scores = df.get('Quality Score', pd.Series('Fair', index=df.index)).fillna('Fair').astype(str).str.strip().str.lower()
            df_valid = df[
                (df['Effect Size'] > 0) &
                (df['Lower CI'] > 0) &
                (df['Upper CI'] > 0) &
                (cases_col >= 50) &
                effect_types.ne('PAF') &
                quality_scores.isin({'good', 'moderate'})
            ].copy()

            if len(df_valid) == 0:
                continue

            # Perform meta-analysis via shared library
            res_dict = meta_analysis.perform_meta_analysis(
                df_valid,
                disease_config['disease_label'],
                folder,
                generate_plots=False,
            )
            headline = res_dict.get('headline')

            if not headline:
                continue

            all_results.append({
                "Exposure": folder,
                "number studies": int(len(df_valid)),
                "Pooled RR": headline.get('pooled_es', 0.0),
                "lower CI RR": headline.get('ci_low', 0.0),
                "upper CI RR": headline.get('ci_upp', 0.0),
                "lower PI RR": headline.get('pi_low'),
                "upper PI RR": headline.get('pi_upp'),
                "I^2 (%)": round(headline.get('i2', 0.0), 1),
                "eggers p-value": headline.get('eggers_p'),
                "total N": int(df_valid['Sample Size'].sum() if 'Sample Size' in df_valid.columns else 0),
                "total Cases": int((df_valid['Cases'].fillna(df_valid.get('Estimated Cases', 0))).sum() if 'Cases' in df_valid.columns else 0)
            })

        except Exception as e:
            print(f"  Error on {folder}: {e}")
            continue

    if not all_results:
        print(f"  No results for {disease_config['disease_label']}.")
        return

    results_df = pd.DataFrame(all_results)
    results_df = results_df.sort_values(by="Pooled RR", ascending=True)

    columns_to_export = [
        "Exposure",
        "number studies",
        "Pooled RR",
        "lower CI RR",
        "upper CI RR",
        "lower PI RR",
        "upper PI RR",
        "I^2 (%)",
        "eggers p-value",
        "total N",
        "total Cases"
    ]
    export_df = results_df[columns_to_export]

    os.makedirs('Plot', exist_ok=True)
    output_file = os.path.join('Plot', disease_config['output_file'])
    export_df.to_excel(output_file, index=False)
    format_workbook(output_file)
    print(f"  OK Exported {len(export_df)} exposures -> {output_file}")


def main():
    for key, config in DISEASES.items():
        run_export(key, config)
    print("\nDone.")


if __name__ == '__main__':
    main()
